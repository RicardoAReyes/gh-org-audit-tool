import requests
import time
import os
import csv
import json
from datetime import datetime
from pathlib import Path
from base64 import b64decode
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# === ORG LIST ===
#list the Github.com organiation names to be processed.
ORG_LIST = ["pypi","nginx","php"]

# === CONFIG ===
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "ghp_your_token_here")
HEADERS = {"Authorization": f"token {GITHUB_TOKEN}"}
REPOS_PER_PAGE = 100
COMMITS_PER_PAGE = 100
DELAY_BETWEEN_REPOS = 1
DELAY_BETWEEN_ORGS = 3
CACHE_FILE = "processed_repos.json"

# === GLOBAL COUNTERS ===
search_request_count = 0

# === Load Repo Cache ===
if Path(CACHE_FILE).exists():
    with open(CACHE_FILE, "r") as f:
        processed_repos = set(json.load(f))
else:
    processed_repos = set()

# === RATE LIMIT HANDLER ===
def handle_rate_limit(request_fn, context, *args, **kwargs):
    response = request_fn(*args, **kwargs)
    if response.status_code != 403:
        return response

    reset_time = response.headers.get("X-RateLimit-Reset")
    remaining = response.headers.get("X-RateLimit-Remaining", "0")
    print(f"  ❌ 403 Forbidden while {context} — Rate limit likely exceeded. Remaining: {remaining}")
    if reset_time:
        wait_seconds = int(reset_time) - int(time.time())
        if wait_seconds > 0:
            print(f"  ⏳ Waiting {wait_seconds} seconds ({round(wait_seconds / 60)} min) until reset...")
            time.sleep(wait_seconds + 1)

    print(f"  🔄 Retrying: {context}")
    retry_response = request_fn(*args, **kwargs)

    if retry_response.status_code != 200:
        print(f"  ❌ Retry failed for {context} — status {retry_response.status_code}")
        with open("failed_requests.log", "a", encoding="utf-8") as log:
            log.write(
                f"{datetime.utcnow().isoformat()} | Retry failed for {context} | Status: {retry_response.status_code}\n"
            )
    return retry_response

# === HELPERS ===

def get_all_repos(org, skipped_repos):
    repos = []
    page = 1
    while True:
        url = f"https://api.github.com/orgs/{org}/repos?per_page={REPOS_PER_PAGE}&page={page}"
        r = handle_rate_limit(lambda: requests.get(url, headers=HEADERS), f"listing repos for {org}")
        if r.status_code != 200:
            skipped_repos.append((org, "[org-level]", f"HTTP {r.status_code} while listing repos"))
            break
        data = r.json()
        if not data:
            break
        repos.extend(data)
        page += 1
        time.sleep(0.5)
    return repos

def get_repo_languages(org, repo_name):
    url = f"https://api.github.com/repos/{org}/{repo_name}/languages"
    r = handle_rate_limit(lambda: requests.get(url, headers=HEADERS), f"getting languages for {org}/{repo_name}")
    if r.status_code != 200:
        return ""
    return ", ".join(list(r.json().keys()))

def search_file_in_repo(org, repo_name, filename, skipped_repos):
    global search_request_count
    search_request_count += 1
    print(f"🔎 [{search_request_count}] Searching {filename} in {org}/{repo_name}")

    search_url = f"https://api.github.com/search/code?q=filename:{filename}+repo:{org}/{repo_name}"
    r = handle_rate_limit(lambda: requests.get(search_url, headers=HEADERS), f"searching {filename} in {org}/{repo_name}")

    time.sleep(2.1)

    if r.status_code != 200:
        return None
    items = r.json().get("items", [])
    return items[0]["path"] if items else None

def get_dockerfile_base_image(org, repo_name, skipped_repos):
    path = search_file_in_repo(org, repo_name, "Dockerfile", skipped_repos)
    if not path:
        return None
    url = f"https://api.github.com/repos/{org}/{repo_name}/contents/{path}"
    r = handle_rate_limit(lambda: requests.get(url, headers=HEADERS), f"fetching Dockerfile content in {org}/{repo_name}")
    if r.status_code != 200:
        return None
    try:
        content = r.json().get("content", "")
        lines = b64decode(content).decode("utf-8").splitlines()
        for line in lines:
            if line.strip().startswith("FROM"):
                return line.strip().split("FROM", 1)[-1].strip()
    except Exception:
        pass
    return None

def get_commit_authors(org, repo_name, skipped_repos):
    authors = {}
    page = 1
    latest_commit = None

    while True:
        url = f"https://api.github.com/repos/{org}/{repo_name}/commits?per_page={COMMITS_PER_PAGE}&page={page}"
        r = handle_rate_limit(lambda: requests.get(url, headers=HEADERS), f"getting commits for {org}/{repo_name}")
        if r.status_code == 409:
            skipped_repos.append((org, repo_name, "No commits"))
            break
        if r.status_code != 200:
            skipped_repos.append((org, repo_name, f"HTTP {r.status_code}"))
            break
        commits = r.json()
        if not commits:
            break
        for commit in commits:
            try:
                name = commit["commit"]["author"]["name"]
                email = commit["commit"]["author"]["email"]
                iso_date = commit["commit"]["author"]["date"]
                date_obj = datetime.strptime(iso_date, "%Y-%m-%dT%H:%M:%SZ")
                if not latest_commit:
                    latest_commit = date_obj.strftime("%m-%d-%Y")
                if name and email and "noreply" not in email:
                    key = (name, email)
                    if key not in authors or date_obj > authors[key]:
                        authors[key] = date_obj
            except Exception:
                continue
        page += 1
        time.sleep(0.5)

    author_rows = [(n, e, d.strftime("%m-%d-%Y")) for (n, e), d in authors.items()]
    return author_rows, latest_commit

def group_rows_by_repo(rows):
    grouped = OrderedDict()
    current_key = None
    buffer = []

    for row in rows:
        if row[0].startswith("Repo URL:"):
            if current_key and buffer:
                grouped[current_key] = buffer
            current_key = row[0]
            buffer = [row]
        else:
            buffer.append(row)

    if current_key and buffer:
        grouped[current_key] = buffer

    return grouped

def write_csv(org, new_rows):
    filename = f"{org.lower()}_contacts.csv"

    existing_rows = []
    if os.path.exists(filename):
        with open(filename, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            existing_rows = list(reader)
    else:
        header = [
            "org", "repo", "repo_status", "languages", "last_commit_date",
            "name", "email", "last_commit_by_user"
        ]

    existing_groups = group_rows_by_repo(existing_rows)
    new_groups = group_rows_by_repo(new_rows)

    merged = OrderedDict(existing_groups)
    merged.update(new_groups)

    with open(filename, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for rows in merged.values():
            writer.writerows(rows)

    print(f"📄 CSV updated: {filename}")

# === MAIN ===

def main():
    global search_request_count
    all_org_data = {}
    skipped_repos = []
    org_summary = []

    for org in ORG_LIST:
        print(f"\n=== 🚀 Processing {org} ===")
        all_rows = []
        repos = get_all_repos(org, skipped_repos)
        repos_scanned = 0
        contacts_found = 0
        dockerfiles_found = 0

        for repo in repos:
            repo_name = repo["name"]
            repo_key = f"{org}/{repo_name}"

            if repo_key in processed_repos:
                print(f"⏩ Skipping already processed repo: {repo_key}")
                continue

            is_archived = repo["archived"]
            repo_status = "archived" if is_archived else "active"
            languages = get_repo_languages(org, repo_name)
            repo_url = f"https://github.com/{org}/{repo_name}"

            docker_base = get_dockerfile_base_image(org, repo_name, skipped_repos)
            if docker_base:
                dockerfiles_found += 1

            if docker_base is None and skipped_repos and "Rate limited" in skipped_repos[-1][-1]:
                print(f"⚠️ Skipping {repo_key} due to Dockerfile rate limit.")
                continue

            authors, last_commit = get_commit_authors(org, repo_name, skipped_repos)
            if not authors:
                continue

            all_rows.append([f"Repo URL: {repo_url}"] + [""] * 7)
            if docker_base:
                all_rows.append([f"Dockerfile | FROM {docker_base}"] + [""] * 7)

            for i, (name, email, date) in enumerate(authors):
                if i == 0:
                    all_rows.append([org, repo_name, repo_status, languages, last_commit, name, email, date])
                else:
                    all_rows.append(["", "", "", "", "", name, email, date])
            all_rows.append([""] * 8)

            contacts_found += len(authors)
            repos_scanned += 1
            processed_repos.add(repo_key)

            with open(CACHE_FILE, "w") as f:
                json.dump(sorted(processed_repos), f, indent=2)

            time.sleep(DELAY_BETWEEN_REPOS)

        write_csv(org, all_rows)
        all_org_data[org] = all_rows
        org_summary.append([org, repos_scanned, len([r for r in skipped_repos if r[0] == org]), contacts_found, dockerfiles_found, search_request_count, len(all_rows)])
        time.sleep(DELAY_BETWEEN_ORGS)

    print("\n📘 Writing Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    # Summary Sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Org Name", "Repos Scanned", "Repos Skipped", "Contacts Found", "Dockerfiles Found", "Search Requests", "CSV Rows"])
    for row in sorted(org_summary, key=lambda x: x[0].lower()):
        summary_ws.append(row)

    # Org Sheets
    for org, rows in all_org_data.items():
        ws = wb.create_sheet(title=org[:31])
        ws.append(["org", "repo", "repo_status", "languages", "last_commit_date", "name", "email", "last_commit_by_user"])
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save("all_orgs_contacts.xlsx")
    print("✅ Excel saved: all_orgs_contacts.xlsx")

    if skipped_repos:
        print("\n⚠️ Skipped Repositories:")
        for org, repo, reason in skipped_repos:
            print(f"- {org}/{repo}: {reason}")
    else:
        print("\n✅ No repos skipped!")

    print(f"\n🔢 Total Dockerfile search requests this run: {search_request_count}")

if __name__ == "__main__":
    main()
